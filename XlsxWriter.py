import openpyxl
import sqlite3
from urllib import parse
import time


def weather_connect():
    return sqlite3.connect('./JMWFinally.db')


def time2date(raw_time_str):
    return time.strftime('%m-%d', time.strptime(raw_time_str, '%Y-%m-%d %H:%M'))


def time2hour(raw_time_str):
    return time.strftime('%H', time.strptime(raw_time_str, '%Y-%m-%d %H:%M'))


def unquote_now(data_tuple):
    ret = list(data_tuple)
    ret[2] = parse.unquote(ret[2])
    ret[4] = parse.unquote(ret[4])
    return ret


def to_map(data_list, unquote_fn):
    """
        用于将形如[("%Y-%m-%d %H:%M",data...)]
        转化为形如{"%m-%d":{"%H":[data...]...}...}
    """
    data_map = {}
    for data_list in list(map(lambda x: unquote_fn(x), list(data_list))):
        date = time2date(data_list[0])
        if date not in data_map.keys():
            data_map[date] = {}
        data_map[date][time2hour(data_list[0])] = data_list[1:]
    return data_map


def get_data_from_database(database_name):
    connection = weather_connect()
    cursor = connection.cursor()
    now_list = cursor.execute(f'SELECT * FROM {database_name}')
    return list(now_list)

alphabet = [chr(i) for i in range(97, 123)]


def to_cell(i_x, i_y):
    # FIXME 丢人!! 列是ABCD,但是A对应0可却没有0行!!也就是说,最顶上的格子的是(0,1)
    return alphabet[i_x] + str(i_y)


def to2str(num):
    if num < 10:
        return "0" + str(num)
    return str(num)


def not_none(table, key, return_list_len):
    if key not in table:
        return ["-" for i in range(0, return_list_len)]
    return table[key]


def fill_sheet(workbook, sheet_name, sheet_index, data_index, data_dict):
    sheet = workbook.create_sheet(sheet_name, sheet_index)
    # 表头
    for i in range(1, 24):
        sheet[to_cell(i, 1)] = i
    # 数据
    y = 2
    for date, value in data_dict.items():
        sheet[to_cell(0, y)] = date
        for hour in range(0, 23):
            x = hour + 1
            sheet[to_cell(x, y)] = not_none(value, to2str(hour), 10)[data_index]
            print(len(value.keys()), hour, date)
        y = y + 1


def fill_sheets(workbook, names_list, data_dict):
    for i in range(0, len(names_list)):
        fill_sheet(workbook, names_list[i], i + 1, i, data_dict)


def fill_now():
    wb = openpyxl.Workbook()
    now_data_dict = to_map(get_data_from_database('now'), unquote_now)
    sheet_names_list = ['气温', '天气', '风向角', '风向', '风力', '风速', '相对湿度', '降水', '能见度']
    fill_sheets(wb, sheet_names_list, now_data_dict)
    wb.save('now.xlsx')


def fill_air():
    wb = openpyxl.Workbook()
    now_data_dict = to_map(get_data_from_database('air'), lambda x: x)
    sheet_names_list = ['空气质量指数AQI', '主要污染物', 'PM10浓度', 'PM2.5浓度', '二氧化氮浓度', '二氧化硫浓度', '一氧化碳浓度', '臭氧浓度']
    fill_sheets(wb, sheet_names_list, now_data_dict)
    wb.save('air.xlsx')


def fill_sun_and_moon_sheets():
    data_list = list(map(lambda x: {time2date(list(x)[0]): list(x)[1:]}, get_data_from_database('sun_and_moon')))
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet('日月升落', 1)
    for y in range(0, len(data_list)):
        sheet[to_cell(0, y + 1)] = list(data_list[y].keys())[0]
        for x in range(0, 4):
            sheet[to_cell(x + 1, y + 1)] = list(data_list[y].values())[0][x]
    wb.save('sun_and_moon.xlsx')

fill_now()
fill_air()
fill_sun_and_moon_sheets()